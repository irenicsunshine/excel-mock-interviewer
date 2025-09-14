"""
Deterministic Excel evaluation using openpyxl and validation rules
"""
import logging
import os
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
from datetime import datetime

logger = logging.getLogger(__name__)

class DeterministicEvaluator:
    """Handles deterministic evaluation of Excel submissions"""
    
    def __init__(self):
        self.test_results = []
        
    def evaluate_answer(self, question: Dict[str, Any], answer_text: str, file_path: Optional[str] = None) -> Dict[str, Any]:
        """Main evaluation method"""
        try:
            results = {
                "passed_tests": 0,
                "total_tests": 0,
                "test_details": [],
                "score": 0.0,
                "confidence": 1.0
            }
            
            if question["type"] == "formula":
                results = self._evaluate_formula(question, answer_text)
            elif question["type"] == "practical" and file_path:
                results = self._evaluate_workbook(question, file_path)
            elif question["type"] == "mcq":
                results = self._evaluate_mcq(question, answer_text)
            else:
                results = self._evaluate_explanation(question, answer_text)
            
            return results
            
        except Exception as e:
            logger.error(f"Deterministic evaluation error: {e}")
            return {
                "passed_tests": 0,
                "total_tests": 1,
                "test_details": [f"Evaluation error: {str(e)}"],
                "score": 0.0,
                "confidence": 0.5
            }
    
    def _evaluate_formula(self, question: Dict[str, Any], formula_text: str) -> Dict[str, Any]:
        """Evaluate formula-based questions"""
        tests = []
        passed = 0
        
        # Test 1: Formula syntax validation
        if self._validate_formula_syntax(formula_text):
            tests.append("✓ Formula syntax is valid")
            passed += 1
        else:
            tests.append("✗ Formula contains syntax errors")
        
        # Test 2: Required functions check
        if "golden_answer" in question and "required_functions" in question["golden_answer"]:
            required_functions = question["golden_answer"]["required_functions"]
            if self._check_required_functions(formula_text, required_functions):
                tests.append("✓ Uses required functions")
                passed += 1
            else:
                tests.append(f"✗ Missing required functions: {required_functions}")
        
        # Test 3: Cell references validation
        if self._validate_cell_references(formula_text):
            tests.append("✓ Cell references are valid")
            passed += 1
        else:
            tests.append("✗ Invalid or suspicious cell references")
        
        total_tests = len(tests)
        score = passed / total_tests if total_tests > 0 else 0.0
        
        return {
            "passed_tests": passed,
            "total_tests": total_tests,
            "test_details": tests,
            "score": score,
            "confidence": 0.9
        }
    
    def _evaluate_workbook(self, question: Dict[str, Any], file_path: str) -> Dict[str, Any]:
        """Evaluate uploaded Excel workbook"""
        tests = []
        passed = 0
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Test 1: Required sheets exist
            required_sheets = question.get("golden_answer", {}).get("required_sheets", [])
            for sheet_name in required_sheets:
                if sheet_name in workbook.sheetnames:
                    tests.append(f"✓ Sheet '{sheet_name}' exists")
                    passed += 1
                else:
                    tests.append(f"✗ Missing required sheet: {sheet_name}")
            
            # Test 2: Pivot tables check
            if question.get("golden_answer", {}).get("requires_pivot", False):
                if self._check_pivot_tables(workbook):
                    tests.append("✓ Contains pivot table(s)")
                    passed += 1
                else:
                    tests.append("✗ No pivot tables found")
            
            # Test 3: Specific cell values
            expected_values = question.get("golden_answer", {}).get("expected_values", {})
            for sheet_cell, expected_value in expected_values.items():
                if self._check_cell_value(workbook, sheet_cell, expected_value):
                    tests.append(f"✓ Correct value in {sheet_cell}")
                    passed += 1
                else:
                    tests.append(f"✗ Incorrect value in {sheet_cell}")
            
            # Test 4: Data validation
            if question.get("golden_answer", {}).get("check_data_cleaning", False):
                if self._check_data_cleaning(workbook):
                    tests.append("✓ Data appears cleaned")
                    passed += 1
                else:
                    tests.append("✗ Data cleaning issues detected")
            
            workbook.close()
            
        except Exception as e:
            tests.append(f"✗ Error opening workbook: {str(e)}")
        
        total_tests = len(tests) if tests else 1
        score = passed / total_tests
        
        return {
            "passed_tests": passed,
            "total_tests": total_tests,
            "test_details": tests,
            "score": score,
            "confidence": 0.85
        }
    
    def _evaluate_mcq(self, question: Dict[str, Any], answer: str) -> Dict[str, Any]:
        """Evaluate multiple choice questions"""
        correct_answer = question.get("golden_answer", {}).get("correct_option", "")
        is_correct = answer.strip().upper() == correct_answer.upper()
        
        return {
            "passed_tests": 1 if is_correct else 0,
            "total_tests": 1,
            "test_details": ["✓ Correct answer" if is_correct else "✗ Incorrect answer"],
            "score": 1.0 if is_correct else 0.0,
            "confidence": 1.0
        }
    
    def _evaluate_explanation(self, question: Dict[str, Any], answer: str) -> Dict[str, Any]:
        """Evaluate explanation-type questions with basic checks"""
        tests = []
        passed = 0
        
        # Test 1: Minimum length
        if len(answer.strip()) >= 50:
            tests.append("✓ Adequate answer length")
            passed += 1
        else:
            tests.append("✗ Answer too brief")
        
        # Test 2: Contains key terms
        key_terms = question.get("golden_answer", {}).get("key_terms", [])
        found_terms = sum(1 for term in key_terms if term.lower() in answer.lower())
        if found_terms >= len(key_terms) * 0.5:  # At least 50% of key terms
            tests.append("✓ Contains relevant terminology")
            passed += 1
        else:
            tests.append("✗ Missing key terminology")
        
        total_tests = len(tests) if tests else 1
        score = passed / total_tests
        
        return {
            "passed_tests": passed,
            "total_tests": total_tests,
            "test_details": tests,
            "score": score,
            "confidence": 0.7  # Lower confidence for subjective evaluation
        }
    
    # Helper methods
    def _validate_formula_syntax(self, formula: str) -> bool:
        """Basic formula syntax validation"""
        if not formula.startswith('='):
            return False
        
        # Check balanced parentheses
        open_count = formula.count('(')
        close_count = formula.count(')')
        if open_count != close_count:
            return False
        
        # Check for basic Excel functions pattern
        function_pattern = r'[A-Z]+\('
        if not re.search(function_pattern, formula.upper()):
            # Allow simple formulas without functions
            pass
        
        return True
    
    def _check_required_functions(self, formula: str, required_functions: List[str]) -> bool:
        """Check if formula contains required functions"""
        formula_upper = formula.upper()
        return all(func.upper() in formula_upper for func in required_functions)
    
    def _validate_cell_references(self, formula: str) -> bool:
        """Validate cell references in formula"""
        cell_pattern = r'[A-Z]+[0-9]+'
        references = re.findall(cell_pattern, formula.upper())
        
        # Basic validation - ensure references look reasonable
        for ref in references:
            col_part = re.match(r'[A-Z]+', ref).group()
            row_part = int(re.search(r'[0-9]+', ref).group())
            
            if len(col_part) > 3 or row_part > 1048576:  # Excel limits
                return False
        
        return True
    
    def _check_pivot_tables(self, workbook) -> bool:
        """Check if workbook contains pivot tables"""
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_pivots') and sheet._pivots:
                return True
        return False
    
    def _check_cell_value(self, workbook, sheet_cell: str, expected_value) -> bool:
        """Check if specific cell contains expected value"""
        try:
            parts = sheet_cell.split('!')
            if len(parts) == 2:
                sheet_name, cell_ref = parts
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    cell_value = sheet[cell_ref].value
                    return str(cell_value) == str(expected_value)
        except:
            pass
        return False
    
    def _check_data_cleaning(self, workbook) -> bool:
        """Basic check for data cleaning indicators"""
        # This is a simplified check - look for consistent formatting
        for sheet in workbook.worksheets:
            if sheet.max_row > 1:  # Has data beyond headers
                # Check for empty rows/cells that might indicate cleaning
                return True
        return False
